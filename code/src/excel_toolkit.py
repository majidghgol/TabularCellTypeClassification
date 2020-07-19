import re
import openpyxl, xlrd
import numpy as np
import string
from nltk.tokenize import word_tokenize
import itertools

col_names = [x for x in string.ascii_lowercase]
col_names += ['{}{}'.format(x,y) for x in string.ascii_lowercase for y in string.ascii_lowercase]
col_names += ['{}{}{}'.format(x,y,z) for x in string.ascii_lowercase for y in string.ascii_lowercase for z in string.ascii_lowercase]

def get_excel_index_as_ints(r):
    col_names = [x for x in string.ascii_lowercase]
    col_names += ['{}{}'.format(x,y) for x in string.ascii_lowercase for y in string.ascii_lowercase]
    col_names += ['{}{}{}'.format(x,y,z) for x in string.ascii_lowercase for y in string.ascii_lowercase for z in string.ascii_lowercase]
    
    if ':' in r:
        rl, rr = r.split(':')
        rl_r, rl_c = int(re.sub('[a-z]', '', rl))-1 , col_names.index(re.sub('[0-9]', '', rl))
        rr_r, rr_c = int(re.sub('[a-z]', '', rr))-1 , col_names.index(re.sub('[0-9]', '', rr))
        i,j = list(range(rl_r, rr_r+1)), list(range(rl_c, rr_c+1))
        ij = list(itertools.product(i,j))
        i = [x[0] for x in ij]
        j = [x[1] for x in ij]
    else:    
        i = [int(re.sub('[a-z]', '', r))-1]
        j = [col_names.index(re.sub('[0-9]', '', r))]
        
    return i,j

def get_excel_range_as_slice(tot_range, block_range):
    col_names = [x for x in string.ascii_lowercase]
    col_names += ['{}{}'.format(x,y) for x in string.ascii_lowercase for y in string.ascii_lowercase]
    col_names += ['{}{}{}'.format(x,y,z) for x in string.ascii_lowercase for y in string.ascii_lowercase for z in string.ascii_lowercase]
    
    
    
    if ':' in tot_range:
        trl, trr = tot_range.split(':')
    else:
        trl = trr = tot_range
        
    tr_tr, tr_br, tr_lc, tr_rc = int(re.sub('[a-z]', '', trl)), int(re.sub('[a-z]', '', trr)), re.sub('[0-9]', '', trl), re.sub('[0-9]', '', trr)
    
    if block_range is None:
        return '{}:{},{}:{}'.format(tr_tr-1, tr_br, col_names.index(tr_lc),col_names.index(tr_rc)+1) 
        
    if ':' in block_range:
        brl, brr = block_range.split(':')
    else:
        brl = brr = block_range
    
    br_tr, br_br, br_lc, br_rc = int(re.sub('[a-z]', '', brl)), int(re.sub('[a-z]', '', brr)), re.sub('[0-9]', '', brl), re.sub('[0-9]', '', brr)
    
    col_slice = '{}:{}'.format(col_names.index(br_lc)-col_names.index(tr_lc), col_names.index(br_rc)-col_names.index(tr_lc)+1)
    row_slice = '{}:{}'.format(br_tr-tr_tr, br_br-tr_tr+1)
    return '{},{}'.format(row_slice, col_slice) 

def get_excel_range_dimension(r):
    col_names = [x for x in string.ascii_lowercase]
    col_names += ['{}{}'.format(x,y) for x in string.ascii_lowercase for y in string.ascii_lowercase]
    col_names += ['{}{}{}'.format(x,y,z) for x in string.ascii_lowercase for y in string.ascii_lowercase for z in string.ascii_lowercase]
    
    if ':' in r:
        rl, rr = r.split(':')
    else:
        return (1,1)
    lrow = int(re.sub('[a-z]', '', rl))
    rrow = int(re.sub('[a-z]', '', rr))
    lcol = col_names.index(re.sub('[0-9]', '', rl))
    rcol = col_names.index(re.sub('[0-9]', '', rr))
    
    return (rrow-lrow+1, rcol-lcol+1)

def get_sheet_names(fpath, file_type='xlsx'):
    res = []
    if file_type == 'xlsx':
        book = openpyxl.load_workbook(fpath, read_only=False, data_only=True)
        sheet_names = book.sheetnames
        book.close()
    else:
        book = xlrd.open_workbook(fpath)
        sheet_names = book.sheet_names()
        book.release_resources()
    return sheet_names

def get_sheet_tarr(fpath, sname, file_type='xlsx', max_cols=100, max_rows=10000):
    res = []
    
    if file_type == 'xlsx':
        book = openpyxl.load_workbook(fpath, read_only=False, data_only=True)
        sheet = book[sname]
        if isinstance(sheet, openpyxl.workbook.workbook.Chartsheet):
            sheet_array = np.array([])
            orig_n, orig_m = 0, 0
        else:
            n, m = sheet.max_row, sheet.max_column
            orig_n, orig_m = n,m
            n = min(n, max_rows)
            m = min(m, max_cols)
            sheet_array = np.empty([n,m], dtype=object)
            for ri, row in enumerate(sheet.values):
                if ri >= n:
                    break
                sheet_array[ri, :] = row[:m]
        book.close()
    else:
        book = xlrd.open_workbook(fpath)
        sheet = book.sheet_by_name(sname)
        n, m = sheet.nrows, sheet.ncols
        orig_n, orig_m = n,m
        n = min(n, max_rows)
        m = min(m, max_cols)
        sheet_array = np.empty([n,m], dtype=object)
        for ri in range(n):
            rvals = sheet.row_values(ri)
            sheet_array[ri] = rvals[:m]
        book.release_resources()
            
    if sheet_array.shape[0] > 0 and sheet_array.shape[1] > 0:
        vf = np.vectorize(lambda x: str(x).strip())
        sheet_array = vf(sheet_array)
        while m > 0 and all([x == '' for x in sheet_array[:,m-1]]):
            m -= 1
        while n > 0 and all([x == '' for x in sheet_array[n-1, :]]):
            n -= 1
        sheet_array = sheet_array[:n, :m]
    else:
        sheet_array = np.array([[]]).astype(str)
    
    return sheet_array, orig_n, orig_m

def get_sheet_cell(fpath, sname, cind, rind, file_type='xlsx'):
    res = []
    
    if file_type == 'xlsx':
        book = openpyxl.load_workbook(fpath, read_only=False, data_only=True)
        sheet = book[sname]
        c = sheet[rind][cind]
        book.close()
    else:
        book = xlrd.open_workbook(fpath)
        sheet = book.sheet_by_name(sname)
        c = sheet[rind][cind]
        book.release_resources()
    return c

def get_book(fname, data_only=False, file_type='xlsx'):
    try:
        if file_type == 'xlsx':
            book = openpyxl.load_workbook(fname, read_only=False, data_only=data_only)
        else:
            book = xlrd.open_workbook(fname, formatting_info=True)
    except Exception as e:
        print(fname)
        raise e
    return book

def get_feature_array(fname, sname, reduced=False, file_type='xlsx'):
    get_fnames = get_feature_names_reduced if reduced else get_feature_names
    book = get_book(fname, data_only=False, file_type=file_type)
    book_data = get_book(fname, data_only=True, file_type=file_type)
    if file_type == 'xlsx':
        sheet = book[sname]
        sheet_data = book_data[sname]
        merged_dict, dependents_dict = get_merged_dict_xlsx(sheet)
        N,M = sheet.max_row, sheet.max_column
        d = len(get_fnames())
        my_cell_features = np.empty([N,M,d], dtype=object)
        for rind, (r_data, r_raw) in enumerate(zip(sheet_data.iter_rows(), sheet.iter_rows())):
            for cind, (c, c_raw) in enumerate(zip(r_data, r_raw)):
                coord = c.coordinate
                rind, cind = get_excel_index_as_ints(coord.lower())
                rind, cind = rind[0], cind[0]
                # rind = c.row - 1
                # cind = c.col_idx - 1
                if coord in dependents_dict:
                    k = dependents_dict[coord]
                    master_cell = sheet_data[k]
                    master_cell_raw = sheet[k]
                    p = get_cell_features_xlsx(master_cell, master_cell_raw, sheet, merged_dict)
                    p['row_ind'] = rind
                    p['col_ind'] = cind
                    p['cell_address'] = coord
                else:
                    p = get_cell_features_xlsx(c, c_raw, sheet, merged_dict)
                ff = properties2features_xlsx(p)

                my_cell_features[rind, cind, :] = [ff[x] for x in get_fnames()]
    else:
        sheet = book.sheet_by_name(sname)
        sheet_data = book_data.sheet_by_name(sname)
        merged_dict, dependents_dict = get_merged_dict_xls(sheet)

        N, M = sheet.nrows, sheet.ncols
        d = len(get_fnames())
        my_cell_features = np.empty([N, M, d], dtype=object)
        my_cell_features[:,:,:] = None
        for rind in range(0, sheet.nrows):
            for cind in range(0, sheet.ncols):
                coord = '{}{}'.format(col_names[cind], rind + 1).upper()
                if coord in dependents_dict:
                    li, lj = dependents_dict[coord]
                    master_cell = sheet_data.cell(li, lj)
                    master_cell_raw = sheet.cell(li, lj)
                    p = get_cell_features_xls(li, lj, master_cell, master_cell_raw, sheet, book, merged_dict)
                    p['row_ind'] = rind
                    p['col_ind'] = cind
                    p['cell_address'] = coord

                else:
                    c = sheet_data.cell(rind, cind)
                    c_raw = sheet.cell(rind, cind)

                    p = get_cell_features_xls(rind, cind, c, c_raw, sheet, book, merged_dict)
                ff = properties2features_xls(p)

                my_cell_features[rind, cind, :] = [ff[x] for x in get_fnames()]
    res = np.array(my_cell_features)
    vfunc = np.vectorize(lambda x: 1 if x == 't' else 0 if x == 'f' else x, otypes='f')
    return vfunc(res)

def get_sheet_annotations(sheet_name, ann_df, n, m):
    annotations = np.empty([n, m], dtype=object)
    annotations[:, :] = None
    indices_array = np.zeros(annotations.shape, dtype=object)
    for i in range(indices_array.shape[0]):
        for j in range(indices_array.shape[1]):
            indices_array[i, j] = (i, j)
    for ri, row in ann_df.iterrows():
        label = row['Annotation.Label']
        if label == 'Table':
            continue
        label = label.lower()
        ann_range = row['Annotation.Range'].replace('$', '')
        sname = row['Sheet.Name']

        if sheet_name == sname:
            l_range = get_excel_range_as_slice(ann_range.lower(), None)
            inds = eval('indices_array[{}]'.format(l_range)).flatten()
            ii, jj = [xx[0] for xx in inds], [xx[1] for xx in inds]
            annotations[(ii, jj)] = label
    return annotations


def properties2features_xlsx(p):
    f = dict()
    f['cell_address'] = p['cell_address']
    f['row_ind'] = p['row_ind']
    f['col_ind'] = p['col_ind']
    f['cell_type=0'] = 't' if p['cell_type'] == 'n' else 'f'
    f['cell_type=2'] = 't' if p['cell_type'] == 'f' else 'f'
    f['length'] = p['length']
    f['words'] = p['words']
    f['leading_spaces'] = p['leading_spaces']
    f['first_char_num'] = p['first_char_num']
    f['first_char_special'] = p['first_char_special']
    f['capitalized'] = p['capitalized']
    f['all_upper'] = p['all_upper']
    f['is_alpha'] = p['is_alpha']
    f['special_chars'] = p['special_chars']
    f['punctuations'] = p['punctuations']
    f['contains_colon'] = p['contains_colon']
    f['words_like_total'] = p['words_like_total']
    f['words_like_table'] = p['words_like_table']
    f['in_year_range'] = p['in_year_range']
    f['is_aggr_formula=1'] = p['is_aggr_formula=1']
    f['ref_val_type=0'] = p['ref_val_type=0']
    f['first_row_num'] = p['first_row_num']
    f['first_col_num'] = p['first_col_num']
    f['num_of_neighbors=0'] = 't' if p['num_of_neighbors'] == 0 else 'f'
    f['num_of_neighbors=1'] = 't' if p['num_of_neighbors'] == 1 else 'f'
    f['num_of_neighbors=2'] = 't' if p['num_of_neighbors'] == 2 else 'f'
    f['num_of_neighbors=3'] = 't' if p['num_of_neighbors'] == 3 else 'f'
    f['num_of_neighbors=4'] = 't' if p['num_of_neighbors'] == 4 else 'f'
    f['h_alignment=0'] = 't' if p['h_alignment'] is None or p['h_alignment'] == 'left' else 'f'
    f['h_alignment=2'] = 't' if p['h_alignment'] == 'center' else 'f'
    f['v_alignment=2'] = 't' if p['v_alignment'] == 'top' else 'f'
    f['indentation'] = p['indentation']
    f['fill_patern=0'] = 't' if p['fill_pattern'] is None else 'f'
    f['is_wraptext'] = 't' if p['is_wraptext'] else 'f'
    f['num_of_cells'] = p['num_of_cells']
    f['border_top_type=0'] = 't' if p['border_top'] is None else 'f'
    f['border_top_type=1'] = 't' if p['border_top'] == 'thin' else 'f'
    f['border_bottom_type=0'] = 't' if p['border_bottom'] is None else 'f'
    f['border_left_type=0'] = 't' if p['border_left'] is None else 'f'
    f['border_right_type=0'] = 't' if p['border_right'] is None else 'f'
    f['border_right_type=2'] = 't' if p['border_right'] == 'medium' else 'f'
    f['cell_borders=0'] = 't' if p['border_top'] is None and p['border_bottom'] is None and p['border_left'] is None and p['border_right'] is None else 'f'
    f['font_height'] = p['font_height']
    f['is_font_color_default'] = 't' if p['font_color'] is None else 'f'
    f['is_bold'] = 't' if p['is_bold'] else 'f'
    f['underline_type=0'] = 't' if p['underline_type'] is None else 'f'
    return f

def properties2features_xls(p):
    f = dict()
    f['cell_address'] = p['cell_address']
    f['row_ind'] = p['row_ind']
    f['col_ind'] = p['col_ind']
    f['cell_type=0'] = 't' if p['cell_type'] == 0 else 'f'
    f['cell_type=2'] = 't' if p['cell_type'] == 2 else 'f'
    f['length'] = p['length']
    f['words'] = p['words']
    f['leading_spaces'] = p['leading_spaces']
    f['first_char_num'] = p['first_char_num']
    f['first_char_special'] = p['first_char_special']
    f['capitalized'] = p['capitalized']
    f['all_upper'] = p['all_upper']
    f['is_alpha'] = p['is_alpha']
    f['special_chars'] = p['special_chars']
    f['punctuations'] = p['punctuations']
    f['contains_colon'] = p['contains_colon']
    f['words_like_total'] = p['words_like_total']
    f['words_like_table'] = p['words_like_table']
    f['in_year_range'] = p['in_year_range']
    f['is_aggr_formula=1'] = p['is_aggr_formula=1']
    f['ref_val_type=0'] = p['ref_val_type=0']
    f['first_row_num'] = p['first_row_num']
    f['first_col_num'] = p['first_col_num']
    f['num_of_neighbors=0'] = 't' if p['num_of_neighbors'] == 0 else 'f'
    f['num_of_neighbors=1'] = 't' if p['num_of_neighbors'] == 1 else 'f'
    f['num_of_neighbors=2'] = 't' if p['num_of_neighbors'] == 2 else 'f'
    f['num_of_neighbors=3'] = 't' if p['num_of_neighbors'] == 3 else 'f'
    f['num_of_neighbors=4'] = 't' if p['num_of_neighbors'] == 4 else 'f'
    f['h_alignment=0'] = 't' if p['h_alignment'] is None or p['h_alignment'] == 0 else 'f'
    f['h_alignment=2'] = 't' if p['h_alignment'] == 2 else 'f'
    f['v_alignment=2'] = 't' if p['v_alignment'] == 2 else 'f'
    f['indentation'] = p['indentation']
    f['fill_patern=0'] = 't' if p['fill_pattern'] == 0 else 'f'
    f['is_wraptext'] = 't' if p['is_wraptext'] != 0 else 'f'
    f['num_of_cells'] = p['num_of_cells']
    f['border_top_type=0'] = 't' if p['border_top'] == 0 else 'f'
    f['border_top_type=1'] = 't' if p['border_top'] == 1 else 'f'
    f['border_bottom_type=0'] = 't' if p['border_bottom'] == 0 else 'f'
    f['border_left_type=0'] = 't' if p['border_left'] == 0 else 'f'
    f['border_right_type=0'] = 't' if p['border_right'] == 0 else 'f'
    f['border_right_type=2'] = 't' if p['border_right'] == 2 else 'f'
    f['cell_borders=0'] = 't' if p['border_top'] == 0 and p['border_bottom'] == 0 and p['border_left'] == 0 and p['border_right'] == 0 else 'f'
    f['font_height'] = p['font_height']
    f['is_font_color_default'] = 't' if p['font_color'] == (255,255,255) else 'f'
    f['is_bold'] = 't' if p['is_bold'] == 1 else 'f'
    f['underline_type=0'] = 't' if p['underline_type'] == 0 else 'f'
    return f

def get_merged_dict_xlsx(sheet):
    merged_cells = list(sheet.merged_cells)
    merged_dict = dict()
    dependents_dict = dict()
    for m in merged_cells:
        left, right = str(m).split(':')
        lcol = col_names.index(re.sub('\d', '', left).lower())
        rcol = col_names.index(re.sub('\d', '', right).lower())

        lrow = int(re.sub('[A-Za-z]', '', left))
        rrow = int(re.sub('[A-Za-z]', '', right))

        dependents = set(['{}{}'.format(col_names[j], i + 1).upper() for i in range(lrow, rrow + 1)
                          for j in range(lcol, rcol + 1) if i != lrow or j != lcol])

        merged_dict[left] = dict(num=(rcol - lcol + 1) * (rrow - lrow + 1),
                                 key=str(m),
                                 left=left,
                                 right=right,
                                 dependents=dependents)
        for x in dependents:
            dependents_dict[x] = left
    return merged_dict, dependents_dict

def get_merged_dict_xls(sheet):
    merged_cells = list(sheet.merged_cells)
    merged_dict = dict()
    dependents_dict = dict()
    for m in merged_cells:
        li, ui, lj, uj = m
        ui -= 1
        uj -= 1
        k = '{}{}:{}{}'.format(col_names[lj], li + 1, col_names[uj], ui + 1).upper()
        left = '{}{}'.format(col_names[lj], li + 1).upper()
        right = (ui, uj)

        dependents = set(['{}{}'.format(col_names[j], i + 1).upper() for i in range(li, ui+1)
                                for j in range(lj,uj+1) if i!=ui or j!=uj])

        merged_dict[left] = dict(num=(uj - lj + 1) * (ui - li + 1),
                                 key=k,
                                 left=left,
                                 right=right,
                                 dependents=dependents)
        for x in dependents:
            dependents_dict[x] = (li,lj)
    return merged_dict, dependents_dict

def get_feature_names_reduced():
    return ['all_upper', 'capitalized',
            'contains_colon', 'first_char_num',
            'first_char_special', 'first_col_num', 'first_row_num',
            'in_year_range',
            'is_alpha',
            'leading_spaces', 'length',
            'num_of_neighbors=0', 'num_of_neighbors=1',
            'num_of_neighbors=2', 'num_of_neighbors=3', 'num_of_neighbors=4',
            'punctuations',
            'special_chars', 'words',
            'words_like_table', 'words_like_total']

def get_feature_names():
    return ['all_upper', 'border_bottom_type=0', 'border_left_type=0',
       'border_right_type=0', 'border_right_type=2', 'border_top_type=0',
       'border_top_type=1', 'capitalized', 'cell_borders=0',
       'cell_type=0', 'cell_type=2', 'contains_colon',
       'fill_patern=0', 'first_char_num',
       'first_char_special', 'first_col_num', 'first_row_num', 'font_height',
       'h_alignment=0', 'h_alignment=2', 'in_year_range',
       'indentation', 'is_aggr_formula=1', 'is_alpha', 'is_bold',
       'is_font_color_default', 'is_wraptext', 'leading_spaces', 'length',
       'num_of_cells', 'num_of_neighbors=0', 'num_of_neighbors=1',
       'num_of_neighbors=2', 'num_of_neighbors=3', 'num_of_neighbors=4',
       'punctuations', 'ref_val_type=0',
       'special_chars', 'underline_type=0', 'v_alignment=2', 'words',
       'words_like_table', 'words_like_total']

def get_cell_features_xlsx(c, c_raw, sheet, merged_dict):
    cell_properties = dict()

    coord = c.coordinate
    rind, cind = get_excel_index_as_ints(coord.lower())
    rind, cind = rind[0], cind[0]
    # rind = c.row - 1
    # cind = c.col_idx - 1
    val_raw = str(c_raw.value)
    val = str(c.value) if c.value else ''
    val_lower = val.lower()
    first_char = val[0] if len(val) > 0 else ''

    cell_properties['row_ind'] = rind
    cell_properties['col_ind'] = cind
    cell_properties['cell_address'] = coord

    cell_properties['cell_type'] = c_raw.data_type
    cell_properties['length'] = len(val)
    cell_properties['words'] = len(word_tokenize(val))
    leading_spaces = 0
    for x in val:
        if x == ' ':
            leading_spaces += 1
        else:
            break
    cell_properties['leading_spaces'] = leading_spaces

    cell_properties['first_char_num'] = 't' if first_char in string.digits else 'f'
    cell_properties[
        'first_char_special'] = 't' if first_char not in string.whitespace + string.ascii_letters + string.digits else 'f'
    cell_properties['capitalized'] = 't' if any([x in string.ascii_uppercase for x in val]) else 'f'
    cell_properties['all_upper'] = 't' if all(
        [x in string.ascii_uppercase + string.whitespace for x in val]) else 'f'
    cell_properties['is_alpha'] = 't' if all(
        [x in string.ascii_letters + string.whitespace for x in val]) else 'f'
    cell_properties['special_chars'] = 't' if any([x not in string.printable for x in val_lower]) else 'f'
    cell_properties['punctuations'] = 't' if any(
        [x in val_lower for x in string.punctuation if x != ':']) else 'f'
    cell_properties['contains_colon'] = 't' if ':' in val_lower else 'f'
    cell_properties['words_like_total'] = 't' if any(
        [x in val_lower for x in ['total', 'average', 'maximum', 'minimum']]) else 'f'
    cell_properties['words_like_table'] = 't' if any([x in val_lower for x in ['table']]) else 'f'
    cell_properties['in_year_range'] = 't' if all(
        [x in string.digits for x in val.strip()]) and val.strip() != '' and 1800 < int(val.strip()) < 2091 else 'f'

    cell_properties['cell_value_raw'] = val_raw
    cell_properties['cell_value'] = val
    cell_properties['ref_val_type=0'] = 'f'
    cell_properties['first_row_num'] = rind
    cell_properties['first_col_num'] = cind

    top_n = sheet.cell(rind, cind + 1) if rind > 0 else None
    bottom_n = sheet.cell(rind + 2, cind + 1)
    right_n = sheet.cell(rind + 1, cind + 2)
    left_n = sheet.cell(rind + 1, cind) if cind > 0 else None
    num_n = sum([1 for x in [top_n, bottom_n, right_n, left_n] if x and x.value != '' and x.value != None])

    cell_properties['num_of_neighbors'] = num_n

    if coord in merged_dict:
        cell_properties['num_of_cells'] = merged_dict[coord]['num']
        cell_properties['cell_address'] = merged_dict[coord]['key']
    else:
        cell_properties['num_of_cells'] = 1
        cell_properties['cell_address'] = coord

    cell_properties['h_alignment'] = c.alignment.horizontal
    cell_properties['v_alignment'] = c.alignment.vertical
    cell_properties['indentation'] = c.alignment.indent
    cell_properties['fill_pattern'] = c.fill.patternType
    cell_properties['is_wraptext'] = 't' if c.alignment.wrapText else 'f'
    cell_properties[
        'is_aggr_formula=1'] = 't' if '=sum(' in val_raw.lower() or '=average(' in val_raw.lower() else 'f'

    if coord in merged_dict:
        c_right = sheet[merged_dict[coord]['right']]
        cell_properties['border_bottom'] = str(c_right.border.bottom.border_style)
        cell_properties['border_right'] = str(c_right.border.right.border_style)

    else:
        cell_properties['border_bottom'] = str(c.border.bottom.border_style)
        cell_properties['border_right'] = str(c.border.right.border_style)

    cell_properties['border_top'] = c.border.top.border_style
    cell_properties['border_left'] = c.border.left.border_style

    cell_properties['font_height'] = c.font.sz
    cell_properties['font_color'] = c.font.color
    cell_properties['is_bold'] = c.font.b
    cell_properties['underline_type'] = c.font.u

    return cell_properties

def get_cell_features_xls(rind, cind, c, c_raw, sheet, book, merged_dict):
    cell_properties = dict()

    N, M = sheet.nrows, sheet.ncols
    coord = '{}{}'.format(col_names[cind], rind + 1).upper()

    val_raw = str(c_raw.value)
    val = str(c.value) if c.value else ''
    val_lower = val.lower()
    first_char = val[0] if len(val) > 0 else ''

    cell_properties['row_ind'] = rind
    cell_properties['col_ind'] = cind
    cell_properties['cell_address'] = coord

    cell_properties['cell_type'] = c_raw.ctype
    cell_properties['length'] = len(val)
    cell_properties['words'] = len(word_tokenize(val))
    leading_spaces = 0
    for x in val:
        if x == ' ':
            leading_spaces += 1
        else:
            break
    cell_properties['leading_spaces'] = leading_spaces

    cell_properties['first_char_num'] = 't' if first_char in string.digits else 'f'
    cell_properties[
        'first_char_special'] = 't' if first_char not in string.whitespace + string.ascii_letters + string.digits else 'f'
    cell_properties['capitalized'] = 't' if any([x in string.ascii_uppercase for x in val]) else 'f'
    cell_properties['all_upper'] = 't' if all(
        [x in string.ascii_uppercase + string.whitespace for x in val]) else 'f'
    cell_properties['is_alpha'] = 't' if all(
        [x in string.ascii_letters + string.whitespace for x in val]) else 'f'
    cell_properties['special_chars'] = 't' if any([x not in string.printable for x in val_lower]) else 'f'
    cell_properties['punctuations'] = 't' if any(
        [x in val_lower for x in string.punctuation if x != ':']) else 'f'
    cell_properties['contains_colon'] = 't' if ':' in val_lower else 'f'
    cell_properties['words_like_total'] = 't' if any(
        [x in val_lower for x in ['total', 'average', 'maximum', 'minimum']]) else 'f'
    cell_properties['words_like_table'] = 't' if any([x in val_lower for x in ['table']]) else 'f'
    cell_properties['in_year_range'] = 't' if all(
        [x in string.digits for x in val.strip()]) and val.strip() != '' and 1800 < int(val.strip()) < 2091 else 'f'

    cell_properties['cell_value_raw'] = val_raw
    cell_properties['cell_value'] = val
    cell_properties['ref_val_type=0'] = 'f'
    cell_properties['first_row_num'] = rind
    cell_properties['first_col_num'] = cind

    top_n = sheet.cell(rind - 1, cind) if rind > 0 else None
    bottom_n = sheet.cell(rind + 1, cind) if rind + 1 < N else None
    right_n = sheet.cell(rind, cind + 1) if cind + 1 < M else None
    left_n = sheet.cell(rind, cind - 1) if cind > 0 else None
    num_n = sum([1 for x in [top_n, bottom_n, right_n, left_n] if x and x.value != '' and x.value != None])

    cell_properties['num_of_neighbors'] = num_n

    if coord in merged_dict:
        cell_properties['num_of_cells'] = merged_dict[coord]['num']
        cell_properties['cell_address'] = merged_dict[coord]['key']
    else:
        cell_properties['num_of_cells'] = 1
        cell_properties['cell_address'] = coord

    if c.xf_index >= len(book.xf_list):
        cell_properties['h_alignment'] = str(0)
        cell_properties['v_alignment'] = str(0)
        cell_properties['indentation'] = str(0)
        cell_properties['fill_pattern'] = str(0)
        cell_properties['is_wraptext'] = 'f'
        cell_properties['is_aggr_furmula=1'] = 't'

        cell_properties['border_top'] = str(0)
        cell_properties['border_bottom'] = str(0)
        cell_properties['border_left'] = str(0)
        cell_properties['border_right'] = str(0)

        cell_properties['font_height'] = 0
        cell_properties['font_color'] = (255, 255, 255)
        cell_properties['is_bold'] = 0
        cell_properties['underline_type'] = 0
        return cell_properties

    xf = book.xf_list[c.xf_index]

    cell_properties['h_alignment'] = str(xf.alignment.hor_align)
    cell_properties['v_alignment'] = str(xf.alignment.vert_align)
    cell_properties['indentation'] = str(xf.alignment.indent_level)
    cell_properties['fill_pattern'] = str(xf.background.fill_pattern)
    cell_properties['is_wraptext'] = 'f' if xf.alignment.text_wrapped == 0 else 't'
    cell_properties['is_aggr_formula=1'] = 't' if c_raw.ctype == 2 else 'f'

    if coord in merged_dict:
        c_right = sheet.cell(merged_dict[coord]['right'][0], merged_dict[coord]['right'][1])
        xf_right = book.xf_list[c_right.xf_index]
        cell_properties['border_bottom'] = str(xf_right.border.bottom_line_style)
        cell_properties['border_right'] = str(xf_right.border.right_line_style)

    else:
        cell_properties['border_bottom'] = str(xf.border.bottom_line_style)
        cell_properties['border_right'] = str(xf.border.right_line_style)

    cell_properties['border_top'] = str(xf.border.top_line_style)
    cell_properties['border_left'] = str(xf.border.left_line_style)

    cell_properties['font_height'] = book.font_list[xf.font_index].height
    cell_properties['font_color'] = book.colour_map.get(book.font_list[xf.font_index].colour_index)
    cell_properties['is_bold'] = book.font_list[xf.font_index].bold
    cell_properties['underline_type'] = book.font_list[xf.font_index].underline_type

    return cell_properties

    